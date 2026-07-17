import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, ScatterChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from collections import Counter
from datetime import datetime

def load_data():
    df = pd.read_csv('data/yes24_bestsellers.csv')
    df['판매가'] = pd.to_numeric(df['판매가'].str.replace(',', ''), errors='coerce')
    df['정가'] = pd.to_numeric(df['정가'].str.replace(',', ''), errors='coerce')
    df['판매지수'] = pd.to_numeric(df['판매지수'].str.replace(',', ''), errors='coerce')
    df['출판일'] = pd.to_datetime(df['출판일'].str.extract(r'(\d{4}년 \d{2}월)')[0], format='%Y년 %m월', errors='coerce')
    df['할인율'] = ((df['정가'] - df['판매가']) / df['정가'] * 100).round(1)
    return df

def extract_keywords(title):
    keywords = re.findall(r'[A-Za-z]+|[\uac00-\ud7a3]{2,}', title)
    return [k.lower() for k in keywords if len(k) > 1]

def create_dashboard(df, wb):
    ws = wb.create_sheet('대시보드', 0)
    ws.sheet_properties.tabColor = '4472C4'
    
    title_font = Font(name='맑은 고딕', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    value_font = Font(name='맑은 고딕', size=12, bold=True)
    
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Yes24 베스트셀러 대시보드'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    stats = [
        ('총 도서 수', f'{len(df):,}권'),
        ('평균 판매가', f'{df["판매가"].mean():,.0f}원'),
        ('평균 할인율', f'{df["할인율"].mean():.1f}%'),
        ('평균 평점', f'{df["평점"].mean():.2f}'),
        ('총 리뷰수', f'{df["리뷰수"].sum():,}개'),
        ('출판사 수', f'{df["출판사"].nunique()}개'),
    ]
    
    for i, (label, value) in enumerate(stats):
        col = i + 1
        ws.cell(row=3, column=col, value=label).font = header_font
        ws.cell(row=3, column=col).fill = header_fill
        ws.cell(row=3, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=3, column=col).border = thin_border
        
        ws.cell(row=4, column=col, value=value).font = value_font
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=4, column=col).border = thin_border
    
    ws.merge_cells('A6:H6')
    ws['A6'] = '출판사별 도서 수 (상위 10개)'
    ws['A6'].font = Font(name='맑은 고딕', size=12, bold=True)
    
    publisher_counts = df['출판사'].value_counts().head(10)
    ws.cell(row=7, column=1, value='출판사')
    ws.cell(row=7, column=2, value='도서 수')
    ws.cell(row=7, column=1).font = header_font
    ws.cell(row=7, column=1).fill = header_fill
    ws.cell(row=7, column=2).font = header_font
    ws.cell(row=7, column=2).fill = header_fill
    
    for i, (pub, count) in enumerate(publisher_counts.items()):
        ws.cell(row=8+i, column=1, value=pub)
        ws.cell(row=8+i, column=2, value=count)
        ws.cell(row=8+i, column=1).border = thin_border
        ws.cell(row=8+i, column=2).border = thin_border
    
    chart1 = BarChart()
    chart1.type = 'col'
    chart1.title = '출판사별 도서 수'
    chart1.y_axis.title = '도서 수'
    chart1.x_axis.title = '출판사'
    chart1.width = 20
    chart1.height = 12
    
    data1 = Reference(ws, min_col=2, min_row=7, max_row=17)
    cats1 = Reference(ws, min_col=1, min_row=8, max_row=17)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.shape = 4
    ws.add_chart(chart1, 'D7')
    
    ws.merge_cells('A20:H20')
    ws['A20'] = '가격 분포'
    ws['A20'].font = Font(name='맑은 고딕', size=12, bold=True)
    
    price_ranges = [
        ('~15,000원', (df['판매가'] <= 15000).sum()),
        ('15,001~20,000원', ((df['판매가'] > 15000) & (df['판매가'] <= 20000)).sum()),
        ('20,001~25,000원', ((df['판매가'] > 20000) & (df['판매가'] <= 25000)).sum()),
        ('25,001~30,000원', ((df['판매가'] > 25000) & (df['판매가'] <= 30000)).sum()),
        ('30,001원~', (df['판매가'] > 30000).sum()),
    ]
    
    ws.cell(row=21, column=1, value='가격대')
    ws.cell(row=21, column=2, value='도서 수')
    ws.cell(row=21, column=1).font = header_font
    ws.cell(row=21, column=1).fill = header_fill
    ws.cell(row=21, column=2).font = header_font
    ws.cell(row=21, column=2).fill = header_fill
    
    for i, (range_name, count) in enumerate(price_ranges):
        ws.cell(row=22+i, column=1, value=range_name)
        ws.cell(row=22+i, column=2, value=count)
        ws.cell(row=22+i, column=1).border = thin_border
        ws.cell(row=22+i, column=2).border = thin_border
    
    chart2 = PieChart()
    chart2.title = '가격 분포'
    chart2.width = 16
    chart2.height = 12
    
    data2 = Reference(ws, min_col=2, min_row=21, max_row=26)
    cats2 = Reference(ws, min_col=1, min_row=22, max_row=26)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.dataLabels = DataLabelList()
    chart2.dataLabels.showPercent = True
    chart2.dataLabels.showVal = True
    ws.add_chart(chart2, 'D20')
    
    ws.merge_cells('A28:H28')
    ws['A28'] = '월별 출판 추이'
    ws['A28'].font = Font(name='맑은 고딕', size=12, bold=True)
    
    monthly = df.groupby(df['출판일'].dt.to_period('M')).size().reset_index(name='도서 수')
    monthly['출판일'] = monthly['출판일'].astype(str)
    
    ws.cell(row=29, column=1, value='출판월')
    ws.cell(row=29, column=2, value='도서 수')
    ws.cell(row=29, column=1).font = header_font
    ws.cell(row=29, column=1).fill = header_fill
    ws.cell(row=29, column=2).font = header_font
    ws.cell(row=29, column=2).fill = header_fill
    
    for i, (_, row) in enumerate(monthly.iterrows()):
        ws.cell(row=30+i, column=1, value=row['출판일'])
        ws.cell(row=30+i, column=2, value=row['도서 수'])
        ws.cell(row=30+i, column=1).border = thin_border
        ws.cell(row=30+i, column=2).border = thin_border
    
    chart3 = LineChart()
    chart3.title = '월별 출판 추이'
    chart3.y_axis.title = '도서 수'
    chart3.width = 20
    chart3.height = 12
    
    data3 = Reference(ws, min_col=2, min_row=29, max_row=29+len(monthly))
    cats3 = Reference(ws, min_col=1, min_row=30, max_row=29+len(monthly))
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    ws.add_chart(chart3, 'D28')
    
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

def create_publisher_sheet(df, wb):
    ws = wb.create_sheet('출판사 분석')
    ws.sheet_properties.tabColor = '70AD47'
    
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    
    ws['A1'] = '출판사명'
    ws['B1'] = '도서 수'
    ws['C1'] = '평균 판매가'
    ws['D1'] = '평균 평점'
    ws['E1'] = '평균 리뷰수'
    ws['F1'] = '평균 할인율'
    
    for col in range(1, 7):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).fill = header_fill
    
    publisher_stats = df.groupby('출판사').agg({
        '도서명': 'count',
        '판매가': 'mean',
        '평점': 'mean',
        '리뷰수': 'mean',
        '할인율': 'mean'
    }).round(1)
    publisher_stats.columns = ['도서 수', '평균 판매가', '평균 평점', '평균 리뷰수', '평균 할인율']
    publisher_stats = publisher_stats.sort_values('도서 수', ascending=False)
    
    for i, (pub, row) in enumerate(publisher_stats.iterrows()):
        ws.cell(row=i+2, column=1, value=pub)
        ws.cell(row=i+2, column=2, value=int(row['도서 수']))
        ws.cell(row=i+2, column=3, value=f'{row["평균 판매가"]:,.0f}원')
        ws.cell(row=i+2, column=4, value=row['평균 평점'])
        ws.cell(row=i+2, column=5, value=f'{row["평균 리뷰수"]:,.0f}개')
        ws.cell(row=i+2, column=6, value=f'{row["평균 할인율"]:.1f}%')
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

def create_price_sheet(df, wb):
    ws = wb.create_sheet('가격 분석')
    ws.sheet_properties.tabColor = 'FFC000'
    
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    ws['A1'] = '도서명'
    ws['B1'] = '출판사'
    ws['C1'] = '판매가'
    ws['D1'] = '정가'
    ws['E1'] = '할인율'
    ws['F1'] = '평점'
    
    for col in range(1, 7):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).fill = header_fill
    
    sorted_df = df.sort_values('판매가', ascending=False)
    
    for i, (_, row) in enumerate(sorted_df.head(50).iterrows()):
        ws.cell(row=i+2, column=1, value=row['도서명'][:30])
        ws.cell(row=i+2, column=2, value=row['출판사'])
        ws.cell(row=i+2, column=3, value=f'{row["판매가"]:,.0f}원')
        ws.cell(row=i+2, column=4, value=f'{row["정가"]:,.0f}원')
        ws.cell(row=i+2, column=5, value=f'{row["할인율"]:.1f}%')
        ws.cell(row=i+2, column=6, value=row['평점'])
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20

def create_popular_sheet(df, wb):
    ws = wb.create_sheet('인기도서')
    ws.sheet_properties.tabColor = 'ED7D31'
    
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    
    ws['A1'] = '순위'
    ws['B1'] = '도서명'
    ws['C1'] = '저자'
    ws['D1'] = '출판사'
    ws['E1'] = '판매지수'
    ws['F1'] = '평점'
    ws['G1'] = '리뷰수'
    
    for col in range(1, 8):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).fill = header_fill
    
    popular_df = df.sort_values('판매지수', ascending=False).head(20)
    
    for i, (_, row) in enumerate(popular_df.iterrows()):
        ws.cell(row=i+2, column=1, value=row['순위'])
        ws.cell(row=i+2, column=2, value=row['도서명'][:40])
        ws.cell(row=i+2, column=3, value=row['저자'][:20])
        ws.cell(row=i+2, column=4, value=row['출판사'])
        ws.cell(row=i+2, column=5, value=f'{row["판매지수"]:,.0f}')
        ws.cell(row=i+2, column=6, value=row['평점'])
        ws.cell(row=i+2, column=7, value=f'{row["리뷰수"]:,.0f}개')
    
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20

def create_keyword_sheet(df, wb):
    ws = wb.create_sheet('키워드 분석')
    ws.sheet_properties.tabColor = '5B9BD5'
    
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    
    ws['A1'] = '키워드'
    ws['B1'] = '출현 횟수'
    
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=2).font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    
    all_keywords = []
    for title in df['도서명']:
        all_keywords.extend(extract_keywords(title))
    
    keyword_counts = Counter(all_keywords).most_common(30)
    
    for i, (keyword, count) in enumerate(keyword_counts):
        ws.cell(row=i+2, column=1, value=keyword)
        ws.cell(row=i+2, column=2, value=count)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

def main():
    print('데이터 로딩 중...')
    df = load_data()
    print(f'총 {len(df)}권의 도서 데이터 로드 완료')
    
    wb = Workbook()
    wb.remove(wb.active)
    
    print('대시보드 시트 생성 중...')
    create_dashboard(df, wb)
    
    print('출판사 분석 시트 생성 중...')
    create_publisher_sheet(df, wb)
    
    print('가격 분석 시트 생성 중...')
    create_price_sheet(df, wb)
    
    print('인기도서 시트 생성 중...')
    create_popular_sheet(df, wb)
    
    print('키워드 분석 시트 생성 중...')
    create_keyword_sheet(df, wb)
    
    print('raw_data 시트 생성 중...')
    ws_raw = wb.create_sheet('raw_data')
    ws_raw.append(list(df.columns))
    for _, row in df.iterrows():
        ws_raw.append(list(row))
    
    output_path = 'data/yes24_dashboard.xlsx'
    wb.save(output_path)
    print(f'대시보드 생성 완료: {output_path}')

if __name__ == '__main__':
    main()
